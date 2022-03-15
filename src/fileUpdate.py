from os import openpty
from random import seed
from turtle import shearfactor
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Color

pastExcelFileName = '/Users/jang-youngjoon/PLAV/PLAV 에셋 /에셋 엑셀파일/구찌(완)/구찌_핸드백.xlsx'
pastExcelFile = openpyxl.load_workbook(pastExcelFileName)
pastExcelFileSheet = pastExcelFile.active

currentExcelFileName = '/Users/jang-youngjoon/PLAV/PLAV 에셋 /에셋 엑셀파일/구찌(완)/구찌 - 20210113.xlsx'
currentExcelFile = openpyxl.load_workbook(currentExcelFileName)
currentExcelFileSheet = currentExcelFile.active

pastExcelFile_max_row_for_a = max((a.row for a in pastExcelFileSheet['A'] if a.value is not None))
currentExcelFile_max_row_for_a = max((a.row for a in currentExcelFileSheet['A'] if a.value is not None))

pastAssetLinkInfo = []
currentAssetLinkInfo = []
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000',fill_type='solid')

# 과거 제품 링크 모아두기
for row_num in range(1, pastExcelFile_max_row_for_a):
    for col_num in range(1,pastExcelFileSheet.max_column):
        if (pastExcelFileSheet.cell(row=row_num, column=col_num).value == "제품 링크"):
            for r in range(1,pastExcelFile_max_row_for_a):
                if (pastExcelFileSheet.cell(row = r, column = col_num).value == None):
                    break;
                # print(pastExcelFileSheet.cell(row = r, column = col_num).value)
                pastAssetLinkInfo.append(pastExcelFileSheet.cell(row = r, column = col_num).value)
            break;



# 현재 제품 링크와 비교
for row_num in range(1, currentExcelFile_max_row_for_a):
    for col_num in range(1,currentExcelFileSheet.max_column):
        if (currentExcelFileSheet.cell(row=row_num, column=col_num).value == "제품 링크"):
            for r in range(1,currentExcelFile_max_row_for_a):
                cellValue = currentExcelFileSheet.cell(row = r, column = col_num).value
                if (cellValue == None):
                    break;
                # 최근 데이터에는 있고, 과거 데이터에는 없는 제품의 열에 red color 칠하기
                if (cellValue not in pastAssetLinkInfo):
                    for cell in currentExcelFileSheet[r]:
                        cell.fill = redFill
            break;


pastExcelFile.save(pastExcelFileName)
currentExcelFile.save(currentExcelFileName)

